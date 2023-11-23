#----importaciones------
import matplotlib.pyplot as plt
import requests
from statistics import mean, median, mode
import pandas as pd
from sympy import (
    Matrix, symbols, Eq, solve, Sum, summation, integrate, 
    integrate, integrate, integrate
)
import numpy as np
import openpyxl
import webbrowser
import sys
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from PIL import Image as PILImage
import shutil
import os
import re
import xlsxwriter
#----importaciones de los modulos----
from modulo_blush import obtener_blushes_powder, obtener_blushes_cream, obtener_blushes_con_calificacion_superior, obtener_blushes_por_marca, obtener_tonos_por_blush
from modulo_eyeshadown import obtener_eyeshadows_en_categoria_palette, obtener_eyeshadows_en_categoria_pencil, obtener_eyeshadows_en_categoria_cream, obtener_eyeshadows_con_calificacion_superior, obtener_eyeshadows_por_marca, obtener_tonos_por_eyeshadow
from modulo_lipstick import obtener_lipsticks_en_categoria_lipstick, obtener_lipsticks_en_categoria_lip_gloss, obtener_lipsticks_en_categoria_liquid, obtener_lipsticks_en_categoria_lip_stain, obtener_lipsticks_con_calificacion_superior, obtener_lipsticks_por_marca, obtener_tonos_por_lipstick
from modulo_mascara import obtener_mascaras_vegan, obtener_mascaras_hipoalergenicas, obtener_mascaras_organicas, obtener_mascaras_con_calificacion_superior, obtener_mascaras_por_marca, obtener_tonos_por_mascara
import modulo_makeup_data

#----- se hace el llamado a la API ------
   #......se crea __main__ para ejecutar el codigo......
if __name__ == "__main__":
    url = "http://makeup-api.herokuapp.com/api/v1/products.json"
    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()
        categorias = ["Blush", "Eyeshadow", "Lipstick"]
        tag = "mascara"
   #....Se crea el llamado al modulo 1 "makeup_data" que contiene el txt y la eliminacion...
      #de datos inecesarios asi como los necesarios
        for category in categorias:
            productos = modulo_makeup_data.get_products_by_category(category)
            if productos:
                filtered_products = modulo_makeup_data.filtro_images(productos)
                modulo_makeup_data.save_to_file("modulo_makeup_info.txt", filtered_products)

        tag_productos = modulo_makeup_data.get_products_by_tag(tag)
        if tag_productos:
            filtro_tag_products = modulo_makeup_data.filtro_images(tag_productos)
            modulo_makeup_data.save_to_file("modulo_makeup_info.txt", filtro_tag_products)
    else:
        print("Error al consultar la API")

#///////////////////// Se comienza con el Menu de Productos y secundarios /////////////////////
  #.....Menu de productos.....
def ProductosOP():
    print("seccion no.1 productos")
    while True:
        print("\n---MENU DE PRODUCTOS---")
        print("1) Blush")
        print("2) Eyeshadown")
        print("3) Lipstick")
        print("4) Mascara")
        print("5) Regresar")
        print("----------------------------------------")
        opcion = input("seleccione una opcion: ")
        if opcion == "1":
            Blush()
        elif opcion == "2":
            Eyeshadown()
        elif opcion == "3":
            Lipstick()
        elif opcion == "4":
            Mascara()
        elif opcion =="5":
            break

#//////////////////////// PRODUCTO BLUSH //////////////////////// 

#------- Se hacen los llamados y creaciones de datos para Categoria Blus -------
  #.....seccion de obtencion de datos de precio para los productos mas caros de blush.....
def obtener_productos_mas_caross(category, product_type):
    url = f"https://makeup-api.herokuapp.com/api/v1/products.json?product_category={category}&product_type={product_type}"
    response = requests.get(url)

    if response.status_code == 200:
        products = response.json()
        sorted_products = sorted(products, key=lambda x: float(x.get('price', 5) or 5), reverse=True)
        return sorted_products[:15] if len(sorted_products) >= 15 else sorted_products[:10]
    else:
        print("Hubo un problema al obtener los productos.")
        return []

def OpcionMasCarosBB():
    categoria = "powder"
    product_type = "blush"
    top_precios_productos = obtener_productos_mas_caross(categoria, product_type)
    if top_precios_productos:
        precios = [float(producto['price']) for producto in top_precios_productos if producto.get('price') is not None]
        for index, producto in enumerate(top_precios_productos, 1):
            print(f"{index}. {producto['name']} - Precio: {producto['price']}")

  #....Calculamos la media, mediana y moda de los precios....
        print("\nOperaciones matematicas:")
        print(f"Media de los precios: {mean(precios)}")
        print(f"Mediana de los precios: {median(precios)}")
        print(f"Moda de los precios: {mode(precios)}")
    else:
        print("No se pudieron obtener los productos de mayor precio.")

def guardar_historial(productos):
    respuesta = input("¿Quieres guardar el historial en un archivo de texto? (si/no): ")
    if respuesta.lower() == "si":
        with open("historial.txt", "a") as archivo:
            for producto in productos:
                archivo.write(str(producto) + "\n")
        print("Historial guardado en el archivo historial.txt")
    else:
        print("El historial no se ha guardado.")
        guardar_historial(OpcionMasCarosBB)

   #....grafico de lineas de los precios mas caros de Blush....
def OpciongraficoB():
    categoria = "powder"
    product_type = "blush"
    top_precios_productos = obtener_productos_mas_caross(categoria, product_type)
    if top_precios_productos:
        precios = [float(producto['price']) for producto in top_precios_productos if producto.get('price') is not None]

        plt.figure(figsize=(10, 6))
        plt.plot(precios, label='5 productos mas caros', marker='o')
        media = mean(precios)
        mediana = median(precios)
        moda = mode(precios)
        plt.axhline(y=media, color='r', linestyle='--', label=f'Media: {media:.2f}')
        plt.axhline(y=mediana, color='g', linestyle='-.', label=f'Mediana: {mediana:.2f}')
        plt.axhline(y=moda, color='b', linestyle=':', label=f'Moda: {moda:.2f}')
        plt.xlabel('indice de productos')
        plt.ylabel('Precio')
        plt.title('Grafico con 15 productos mas caros de blush')
        plt.legend()
        plt.grid(True)
        plt.show()
        
  #.....seccion de obtencion de datos de precio para los productos mas baratos de blush.....

def obtener_productosbaratos(category, product_type):
    url = f"https://makeup-api.herokuapp.com/api/v1/products.json?product_category={category}&product_type={product_type}"
    response = requests.get(url)

    if response.status_code == 200:
        products = response.json()
        sorted_products = sorted(products, key=lambda x: float(x.get('price', 0) or 0))
        return sorted_products[:15] 
    else:
        print("Hubo un problema al obtener los productos.")
        return []

def OpcionMasBaratosBB():
    categoria = "powder"
    product_type = "blush"
    productos_mas_baratos = obtener_productosbaratos(categoria, product_type)
    if productos_mas_baratos:
        precios = [float(producto['price']) for producto in productos_mas_baratos if producto.get('price') is not None]
        for index, producto in enumerate(productos_mas_baratos, 1):
            print(f"{index}. {producto['name']} - Precio: {producto['price']}")
        #....Calculamos la media, mediana y moda de los precios....
        if len(precios) > 0:
            print("\nOperaciones matematicas:")
            print(f"Media de los precios: {mean(precios)}")
            print(f"Mediana de los precios: {median(precios)}")
            print(f"Moda de los precios: {mode(precios)}")
        else:
            print("No se pudieron calcular las estadisticas, precios no disponibles.")
    else:
        print("No se pudieron obtener los productos de menor precio.")
#....grafico de lineas de los precios mas baratos de Blush....
def opciongraficoBB():
    categoria = "powder"
    product_type = "blush"
    productos_mas_baratos = obtener_productosbaratos(categoria, product_type)
    if productos_mas_baratos:
        precios = [float(producto['price']) for producto in productos_mas_baratos if producto.get('price') is not None]

        plt.figure(figsize=(10, 6))
        colores = ['#BF00BF', '#00BFBF', '#BFBF00', '#DC143C']

        plt.plot(precios, label='15 productos mas baratos', marker='o', color=colores[0])
        media = mean(precios)
        mediana = median(precios)
        moda = mode(precios)
        plt.axhline(y=media, color=colores[1], linestyle='--', label=f'Media: {media:.2f}')
        plt.axhline(y=mediana, color=colores[2], linestyle='-.', label=f'Mediana: {mediana:.2f}')
        plt.axhline(y=moda, color=colores[3], linestyle=':', label=f'Moda: {moda:.2f}')
        plt.xlabel('indice de productos')
        plt.ylabel('Precio')
        plt.title('Grafico con 15 productos mas baratos de blush')
        plt.legend()
        plt.grid(True)
        plt.show()


#grafico de marcas 
def obtener_grafico_blushes_por_marca():
    url = "https://makeup-api.herokuapp.com/api/v1/products.json?product_type=blush"

    try:
        response = requests.get(url)

        if response.status_code == 200:
            datos = response.json()
            blushes_por_marca = {}

            for producto in datos:
                marca = producto.get('brand', 'Marca no disponible')

                if marca in blushes_por_marca:
                    blushes_por_marca[marca] += 1
                else:
                    blushes_por_marca[marca] = 1

            # Crear el gráfico de pastel
            labels = blushes_por_marca.keys()
            sizes = blushes_por_marca.values()

            plt.figure(figsize=(8, 8))
            plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
            plt.axis('equal')
            plt.title('Cantidad de blushes por marca')
            plt.show()

        else:
            print(f"Error al obtener datos. Codigo de estado: {response.status_code}")

    except requests.RequestException as e:
        print(f"Error de solicitud: {e}")

#...graficod de tonos de blush...
def grafico_tonos_por_blush():
    url = "https://makeup-api.herokuapp.com/api/v1/products.json?product_type=blush"

    try:
        response = requests.get(url)

        if response.status_code == 200:
            datos = response.json()
            tonos_por_blush = {}
            
            for producto in datos:
                nombre = producto.get('name', 'Nombre no disponible')
                tonos = producto.get('product_colors', [])
                tonos_por_blush[nombre] = tonos
            
            # Extraer los nombres de los tonos y asignarlos a un conjunto para evitar duplicados
            nombres_tonos = set()
            for tonos in tonos_por_blush.values():
                for tono in tonos:
                    nombre_tono = tono.get('colour_name', 'Nombre no disponible')
                    nombres_tonos.add(nombre_tono)
            
            # Crear un diccionario para asignar un numero a cada tono
            tonos_numerados = {nombre_tono: i+1 for i, nombre_tono in enumerate(nombres_tonos)}

            # Crear listas para el grafico de dispersion
            nombres_blush = []
            numeros_tonos = []

            for nombre, tonos in tonos_por_blush.items():
                nombres_blush.extend([nombre] * len(tonos))
                numeros_tonos.extend([tonos_numerados.get(t.get('colour_name')) for t in tonos])

            # Crear el grafico de dispersion con el color rosa
            plt.figure(figsize=(10, 6))
            plt.scatter(nombres_blush, numeros_tonos, color='deeppink', alpha=0.5)
            plt.xlabel('Blush')
            plt.ylabel('Cantidad Tonos de Blush')
            plt.title('Grafico de tonos de Blush')
            plt.xticks(rotation=90)
            plt.tight_layout()
            plt.show()

        else:
            print(f"Error al obtener datos. Código de estado: {response.status_code}")

    except requests.RequestException as e:
        print(f"Error de solicitud: {e}")


#-------------------Menu de Productos Blush------------------
def Blush():
    while True:
        print("\n---MENU DE BLUSH---")
        print("1) Categoría")
        print("2) Mejores calificaciones")
        print("3) Precio")
        print("4) Marcas")
        print("5) Tonos")
        print("6) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opción: ")
        if elegirOpcion == "1":
            CategoriaBlush()
        elif elegirOpcion == "2":
            print("Mejores calificaciones")
            mejor_Calificados = obtener_blushes_con_calificacion_superior(4)
            print(mejor_Calificados)
        elif elegirOpcion == "3":
            PrecioBlush()
        elif elegirOpcion == "4":
            MarcasBlush()           
        elif elegirOpcion == "5":
            TonosBlush()
        elif elegirOpcion == "6":
            break
  #.......Menu categorias de Blush.....
def CategoriaBlush():
    while True:
        print("\n---CATEGORIA DE BLUSH---")
        print("1) En Polvo")
        print("2) En Crema")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opción: ")
        if elegirOpcion == "1":
            print("En Polvo")
            powder = obtener_blushes_powder()
            print(powder)
        elif elegirOpcion == "2":
            print("En Crema")
            cream = obtener_blushes_cream()
            print(cream)
        elif elegirOpcion == "3":
            break

  #.......Menu Marcas de Blush.....
def MarcasBlush():
    while True:
        print("\n---MARCAS BLUSH---")
        print("1) Informacion de marcas")
        print("2) Ver grafico")
        print("3) Salir")
        print("--------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Las Marcas de Blush son: ")
            marcas = obtener_blushes_por_marca()
            print(marcas)
        elif elegirOpcion == "2":
            print(" ver grafico: ")
            obtener_grafico_blushes_por_marca()
        elif elegirOpcion == "3":
            break

  #.......Menu Tonos de Blush.....
def TonosBlush():
    while True:
        print("\n---TONOS DE BLUSH---")
        print("1) Tonos de marcas")
        print("2) Ver grafico")
        print("3) Salir")
        print("--------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Los Tonos de Blush son: ")
            tonos = obtener_tonos_por_blush()
            print(tonos)
        elif elegirOpcion == "2":
            print(" ver grafico: ")
            grafico_tonos_por_blush()
        elif elegirOpcion == "3":
            break

  #.......Menu Precios de Blush.....
def PrecioBlush():
    while True:
        print("\n---PRECIO DE BLUSH---")
        print("1) Mas caros")
        print("2) Mas baratos")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            OpcionMasCarosB()
        elif elegirOpcion == "2":
            OpcionMasBaratosB()
        elif elegirOpcion == "3":
            break
    #________Menu Precios mas caros de Blush______
def OpcionMasCarosB():
    while True:
        print("\n---MAS CAROS---")
        print("1) Ver productos")
        print("2) Ver grafico")
        print("3) Historial")
        print("4) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Los 15 productos mas caros de Blush son: ")
            OpcionMasCarosBB()
        elif elegirOpcion == "2":
            print("Ver grafico")
            OpciongraficoB()
        elif elegirOpcion == "3":
            print("El Historial es:")
            guardar_historial(productos)
        elif elegirOpcion == "4":
            break

    #________Menu Precios mas baratos de Blush______
def OpcionMasBaratosB():
    while True:
        print("\n---MAS BARATOS---")
        print("1) Ver productos")
        print("2) Ver grafico")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Los 15 productos mas baratos de Blush son: ")
            OpcionMasBaratosBB()
        elif elegirOpcion == "2":
            print("Ver grafico")
            opciongraficoBB()
        elif elegirOpcion == "3":
            break

#//////////////////////// PRODUCTO EYESHADOWN //////////////////////// 
#------- Se hacen los llamados y creaciones de datos para Categoria Eyeshadown -------
  
#.....seccion de obtencion de datos de precio para los productos mas caros de Eyeshadow.....
def obtener_productos_mas_carosE(categories):
    top_precios = []
    for category in categories:
        url = f"https://makeup-api.herokuapp.com/api/v1/products.json?product_category={category}&product_type=eyeshadow"
        response = requests.get(url)
        if response.status_code == 200:
            products = response.json()
            top_precios.extend(products)
        else:
            print(f"Hubo un problema al obtener los productos de la categoria {category}.")
    return top_precios

def OpmascarosE():
    categorias = ["palette", "pencil", "cream"]
    top_precios_productos = obtener_productos_mas_carosE(categorias)
    precios = []  
    if top_precios_productos:
        top_precios_productos = [producto for producto in top_precios_productos if producto.get('price') is not None]
        precios = [float(producto['price']) for producto in top_precios_productos]  
        top_precios_productos = sorted(top_precios_productos, key=lambda x: float(x.get('price', 0)), reverse=True)[:15]
        
        for index, producto in enumerate(top_precios_productos, 1):
            print(f"{index}. {producto['name']} - Precio: {producto['price']}") 

            #.......Operacion Matematica......
            rango_precios = max(precios) - min(precios)
        print(" ")
        print("Operacion Matematica:")
        print(f"El rango de precios de los 15 productos mas caros es: {rango_precios}") 
    return precios 
   #.....Grafico de barras para los precios de los productos mas caros de Eyeshadow.....
def graficoE(precios, rango):
    plt.figure(figsize=(10, 6))
    plt.bar(range(len(precios)), precios, color='#DC143C', alpha=0.7, label='Precios')
    # Linea horizontal para el rango de precios
    plt.axhline(rango, color='#008B8B', linestyle='--', label=f'Rango de precios: {rango}')
    plt.xlabel('Productos')
    plt.ylabel('Precio')
    plt.title('Precios de los 15 productos mas caros de Eyeshadow')
    plt.legend()
    plt.xticks(range(len(precios)), [f'Producto {i+1}' for i in range(len(precios))], rotation=45)
    plt.tight_layout()
    plt.show()

#.....seccion de obtencion de datos de precio para los productos mas baratos de Eyeshadow.....

def obtener_productos_mas_baratosE(categories):
    cheapest_products = []
    for category in categories:
        url = f"https://makeup-api.herokuapp.com/api/v1/products.json?product_category={category}&product_type=eyeshadow"
        response = requests.get(url)
        if response.status_code == 200:
            products = response.json()
            cheapest_products.extend(products)
        else:
            print(f"Hubo un problema al obtener los productos de la categoria {category}.")
    return cheapest_products

def obtener_mas_baratos_eyeshadow():
    categories = ["palette", "pencil", "cream"]
    cheapest_products_eyeshadow = obtener_productos_mas_baratosE(categories)
    cheapest_eyeshadows = []     
    if cheapest_products_eyeshadow:
        cheapest_eyeshadows = [producto for producto in cheapest_products_eyeshadow if producto.get('price') is not None]
        cheapest_eyeshadows = sorted(cheapest_eyeshadows, key=lambda x: float(x.get('price', 0)))[:15]   
        return cheapest_eyeshadows
    else:
        return []

def mostrar_eyeshadows_mas_baratos():
    cheapest_eyeshadows = obtener_mas_baratos_eyeshadow()   
    if cheapest_eyeshadows:
        print("Los 15 productos mas baratos de Eyeshadow son:")
        for index, producto in enumerate(cheapest_eyeshadows, 1):
            price = float(producto['price']) if producto['price'] else None
            print(f"{index}. {producto['name']} - Precio: {price}")
        
        #.........Realizar operacion matematica con los precios mas baratos......
        precios = [float(producto['price']) for producto in cheapest_eyeshadows if producto['price']]
        suma_precios = sum(precios)
        print(f"\nLa suma de los precios de los 15 productos mas baratos es: {suma_precios}")
    else:
        print("No se encontraron productos de sombras de ojos mas baratos.")

   #.........Realizar grafico de barras con los precios mas baratos......
def generar_grafico_eyeshadows():
    cheapest_eyeshadows = obtener_mas_baratos_eyeshadow()   
    if cheapest_eyeshadows:
        nombres_productos = [producto['name'] for producto in cheapest_eyeshadows]
        precios_productos = [float(producto['price']) if producto['price'] else None for producto in cheapest_eyeshadows]
        plt.figure(figsize=(10, 6))
        plt.barh(nombres_productos, precios_productos, color='#BFBF00')
        plt.xlabel('Precio')
        plt.title('Precios de los 15 productos mas baratos de Eyeshadow')
        plt.tight_layout()
        plt.show()
    else:
        print("No se encontraron productos de sombras de ojos mas baratos.")

#.....grafico de marcas de eyeshadow....
def grafica_marcas_eyeshadow(marcas, conteos):
    plt.pie(conteos, labels=marcas, autopct='%1.1f%%')
    plt.axis('equal')  
    plt.title('Distribucion de marcas de Eyeshadow')
    plt.show()

url = "https://makeup-api.herokuapp.com/api/v1/products.json"
params = {"product_type": "eyeshadow"}
response = requests.get(url, params=params)

if response.status_code == 200:
    data = response.json()
    marcas = [producto['brand'] for producto in data]
    conteo_marcas = {}
    for marca in marcas:
        conteo_marcas[marca] = conteo_marcas.get(marca, 0) + 1

    marcas_unicas = list(conteo_marcas.keys())
    cantidad_por_marca = list(conteo_marcas.values())


#-------------------Menu de Productos Eyeshadown------------------
def Eyeshadown():
  while True:
    print("\n---MENU DE EYESHADOWN---")
    print("1) Categoria")
    print("2) Mejores calificaciones")
    print("3) precio")
    print("4) Marcas")
    print("5) Regresar")
    print("----------------------------------------")
    elegirOpcion = input("elige una opcion: ")
    if elegirOpcion == "1":
      CategoriaEyeshadown()
    elif elegirOpcion == "2":
      print("Mejores calificaciones")
      mejores_eyeshadown = obtener_blushes_con_calificacion_superior(4)
      print(mejores_eyeshadown)
    elif elegirOpcion == "3":
      PrecioEyeshadown()
    elif elegirOpcion == "4":
      print("Marcas")
      MarcasEyeshadow()   
    elif elegirOpcion == "5":
      break

  #.......Menu categorias de Eyeshadown.....
def CategoriaEyeshadown():
    while True:
        print("\n---CATEGORIA DE EYESHADOWN---")
        print("1) Paletas de Colores")
        print("2) En Lapiz")
        print("3) En Crema")
        print("4) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Paletas de Colores")
            palette = obtener_eyeshadows_en_categoria_palette()
            print(palette)
        elif elegirOpcion == "2":
            print("En Lapiz")
            pencil = obtener_eyeshadows_en_categoria_pencil()
            print(pencil)
        elif elegirOpcion == "3":
            print("En Crema")  
            enCrema = obtener_eyeshadows_en_categoria_cream()
            print(enCrema)  
        elif elegirOpcion == "4":
            break

  #.......Menu Marcas de Eyeshadow.....
def MarcasEyeshadow():
    while True:
        print("\n---MARCAS EYESHADOW---")
        print("1) Informacion de marcas")
        print("2) Ver grafico")
        print("3) Salir")
        print("--------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Las Marcas de Eyeshadow son: ")
            marcasEyeshadown = obtener_eyeshadows_por_marca()
            print(marcasEyeshadown)
        elif elegirOpcion == "2":
            print(" ver grafico: ")
            grafica_marcas_eyeshadow(marcas_unicas, cantidad_por_marca)
        elif elegirOpcion == "3":
            break
    
  #.......Menu precios de Eyeshadown.....
def PrecioEyeshadown():
    while True:
        print("\n---PRECIO DE EYESHADOWN---")
        print("1) Mas caros")
        print("2) Mas baratos")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            OpcionMasCarosE()
        elif elegirOpcion == "2":
            OpcionMasBaratosE()
        elif elegirOpcion == "3":
            break

    #________Menu Precios mas caros de Eyeshadown______
def OpcionMasCarosE():
    precios = []  
    while True:
        print("\n---MAS CAROS---")
        print("1) Ver productos")
        print("2) Ver grafico")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("Elige una opcion: ")
        if elegirOpcion == "1":
            print("Los 15 productos mas caros de Eyeshadow son: ")
            precios = OpmascarosE()  
        elif elegirOpcion == "2":
            if not precios:  
                print("Primero obten los datos en la opcion 1.")
            else:
                rango = max(precios) - min(precios)
                graficoE(precios, rango)          
        elif elegirOpcion == "3":
            break

    #________Menu Precios mas baratos de Eyeshadown______
def OpcionMasBaratosE():
    while True:
        print("\n---MAS BARATOS---")
        print("1) Ver productos")
        print("2) Ver grafico")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            mostrar_eyeshadows_mas_baratos()
        elif elegirOpcion == "2":
            generar_grafico_eyeshadows()
        elif elegirOpcion == "3":
            break

#//////////////////////// PRODUCTO LIPSTICK //////////////////////// 

#------- Se hacen los llamados y creaciones de datos para Categoria Lipstick -------
  #.....seccion de obtencion de datos de precio para los productos Lipstick.....

  #.......obtencion de los productos mas caros de lipstick......

def obtener_productos_mas_carosL(categories):
    expensive_products = []
    for category in categories:
        url = f"https://makeup-api.herokuapp.com/api/v1/products.json?product_category={category}&product_type=lipstick"
        response = requests.get(url)
        if response.status_code == 200:
            products = response.json()
            expensive_products.extend(products)
        else:
            print(f"Hubo un problema al obtener los productos de la categoria {category}.")
    return expensive_products

def obtener_mas_caros_lipsticks():
    categories = ["lipstick", "lip_gloss", "liquid", "lip_stain"]
    expensive_lipsticks = obtener_productos_mas_carosL(categories)
    most_expensive_lipsticks = []

    if expensive_lipsticks:
        expensive_lipsticks = [producto for producto in expensive_lipsticks if producto.get('price') is not None]
        most_expensive_lipsticks = sorted(expensive_lipsticks, key=lambda x: float(x.get('price', 0)), reverse=True)[:15]
        
    return most_expensive_lipsticks

def mostrar_lipsticks_mas_caros():
    most_expensive_lipsticks = obtener_mas_caros_lipsticks()
    
    if most_expensive_lipsticks:
        print("Los 15 productos mas caros de lipsticks son:")
        for index, producto in enumerate(most_expensive_lipsticks, 1):
            price = float(producto['price']) if producto['price'] else None
            print(f"{index}. {producto['name']} - Precio: {price}")

        precios = [float(producto['price']) for producto in most_expensive_lipsticks]
        
        #.......Operacion matematica.....
        #La integral calculada mostrara el resultado en terminos de la variable 'x'. 
        #Esta expresion es una representacion simbolica de la integral, donde se muestra la formula resultante en funcion de 'x'.
        x = symbols('x')
        funcion_polinomica = precios[0] * x**3 + precios[1] * x**2 + precios[2] * x + precios[3]
        integral = integrate(funcion_polinomica, x)
        print(" ")
        print("Operacion matematica: ")
        print(f"La integral de la funcion de precios es: {integral}")
        max_price = max(precios)
        #.....Calculamos el valor numerico de la integral en el rango del precio mas alto.....
        valor_numerico_integral = integral.subs(x, max_price) - integral.subs(x, 0)
        print(f"El valor numerico de la integral en el rango del precio mas alto ({max_price}) es: {valor_numerico_integral}")
        return most_expensive_lipsticks, precios, integral, max_price, valor_numerico_integral
 #....Grafico de linea de los precios mas caros de lipstick....
def parq():
    most_expensive_lipsticks, precios, integral, max_price, valor_numerico_integral = mostrar_lipsticks_mas_caros()
    x = np.linspace(0, max(precios), 100)
    y = precios[0] * x**3 + precios[1] * x**2 + precios[2] * x + precios[3]
    plt.figure(figsize=(8, 6))
    plt.plot(x, y, label='Funcion de precios')
    plt.scatter(max_price, precios[0] * max_price**3 + precios[1] * max_price**2 + precios[2] * max_price + precios[3], color='red', label=f"Precio mas alto: {max_price}")
    plt.xlabel('Precios')
    plt.ylabel('Valor de la funcion')
    plt.title('Funcion de precios mas caros de lipstick y punto mas alto ')
    plt.legend()
    plt.show()


  #.......obtencion de los productos mas baratos de lipstick......

def obtener_productos_mas_baratosL(categories):
    all_products = []

    for category in categories:
        url = f"https://makeup-api.herokuapp.com/api/v1/products.json?product_category={category}&product_type=lipstick"
        response = requests.get(url)
        
        if response.status_code == 200:
            products = response.json()
            all_products.extend(products)
        else:
            print(f"Hubo un problema al obtener los productos de la categoria {category}.")

    valid_products = [product for product in all_products if product.get('price') and isinstance(product['price'], (int, float, str))]
    sorted_products = sorted(valid_products, key=lambda x: float(x['price']))[:15]
    
    return sorted_products

def mostrar_productos_mas_baratos_lipstick():
    productos_mas_baratos = obtener_productos_mas_baratosL(["lipstick", "lip_gloss", "liquid", "lip_stain"])
    if productos_mas_baratos:
        print("Los 15 productos mas baratos de lipstick son:")
        for index, producto in enumerate(productos_mas_baratos, 1):
            price = float(producto['price']) if producto['price'] else None
            print(f"{index}. {producto['name']} - Precio: {price}")
     #.....Creacion de la operacion matematica lanzando el minimo y maximo de los precios
        precios = [float(producto['price']) for producto in productos_mas_baratos]
        precio_maximo = max(precios)
        precio_minimo = min(precios)
        operacion_matematica = precio_maximo - precio_minimo

        print(f"Operacion matematica: Diferencia entre el precio maximo y minimo: {operacion_matematica}")
    #....Grafico de barras para los 15 productos mas baratos
def mostrar_grafica_diferencia_precioL():
    productos_mas_baratos = obtener_productos_mas_baratosL(["lipstick", "lip_gloss", "liquid", "lip_stain"])

    if productos_mas_baratos:
        nombres_productos = [producto['name'] for producto in productos_mas_baratos]
        precios = [float(producto['price']) for producto in productos_mas_baratos]
        precio_maximo = max(precios)
        precio_minimo = min(precios)
        operacion_matematica = precio_maximo - precio_minimo
        plt.figure(figsize=(10, 6))
        plt.barh(nombres_productos, precios, color='skyblue')
        plt.xlabel('Precios')
        plt.title('Precios de los 15 productos mas baratos de lipstick')
        plt.axhline(y=operacion_matematica, color='red', linestyle='--', label='Diferencia Precio Maximo y Minimo')
        plt.legend()
        plt.tight_layout()
        plt.show()

#....grafico de marcas de lipstick....
def grafica_marcas_lipstick(marcas, conteos):
    plt.pie(conteos, labels=marcas, autopct='%1.1f%%')
    plt.axis('equal')  
    plt.title('Distribucion de marcas de Lipstick')
    plt.show()


url = "https://makeup-api.herokuapp.com/api/v1/products.json"
params = {"product_type": "lipstick"}
response = requests.get(url, params=params)

if response.status_code == 200:
    data = response.json()
    marcas = [producto['brand'] for producto in data]
    conteo_marcas = {}
    for marca in marcas:
        conteo_marcas[marca] = conteo_marcas.get(marca, 0) + 1

    marcas_unicas = list(conteo_marcas.keys())
    cantidad_por_marca = list(conteo_marcas.values())

#....grafico tonos lipstick.....

def grafico_tonos_lipstick():
    url = "https://makeup-api.herokuapp.com/api/v1/products.json?product_type=lipstick"

    try:
        response = requests.get(url)

        if response.status_code == 200:
            datos = response.json()
            tonos_lipstick = {}

            for producto in datos:
                nombre = producto.get('name', 'Nombre no disponible')
                tonos = producto.get('product_colors', [])
                tonos_lipstick[nombre] = tonos
            
            nombres_tonos = set()
            for tonos in tonos_lipstick.values():
                for tono in tonos:
                    nombre_tono = tono.get('colour_name', 'Nombre no disponible')
                    nombres_tonos.add(nombre_tono)
            
            # Crear un diccionario para asignar un numero a cada tono
            tonos_numerados = {nombre_tono: i+1 for i, nombre_tono in enumerate(nombres_tonos)}

            # Crear listas para el grafico de dispersion
            nombres_lipstick = []
            numeros_tonos = []

            for nombre, tonos in tonos_lipstick.items():
                nombres_lipstick.extend([nombre] * len(tonos))
                numeros_tonos.extend([tonos_numerados.get(t.get('colour_name')) for t in tonos])

            # Crear el grafico de dispersion
            plt.figure(figsize=(10, 6))
            plt.scatter(nombres_lipstick, numeros_tonos, color='red', alpha=0.5)
            plt.xlabel('Labiales')
            plt.ylabel('Cantidad de Tonos de Labial')
            plt.title('Grafico de dispersion de tonos de Labial')
            plt.xticks(rotation=90)
            plt.tight_layout()
            plt.show()

        else:
            print(f"Error al obtener datos. Codigo de estado: {response.status_code}")

    except requests.RequestException as e:
        print(f"Error de solicitud: {e}")



#-------------------Menu de Productos Lipstick------------------
def Lipstick():
  while True:
    print("\n---MENU DE LIPSTICK---")
    print("1) Categoria")
    print("2) Mejores calificaciones")
    print("3) precio")
    print("4) Marcas")
    print("5) Tonos")
    print("6) Regresar")
    print("----------------------------------------")
    elegirOpcion = input("elige una opcion: ")
    if elegirOpcion == "1":
      CategoriaLipstick()
    elif elegirOpcion == "2":
      print("Mejores calificaciones")
      mejores_lipsticks = obtener_lipsticks_con_calificacion_superior(4)
      print(mejores_lipsticks)
    elif elegirOpcion == "3":
      PrecioLipstick()
    elif elegirOpcion == "4":
      print("Marcas")
      MarcasLipstick()
    elif elegirOpcion == "5":
      print("Tonos")
      TonosEyeshadow()
    elif elegirOpcion == "6":
      break

  #.......Menu categorias de Lipstick.....
def CategoriaLipstick():
    while True:
        print("\n---CATEGORIA DE LIPSTICK---")
        print("1) Lipstick")
        print("2) Brillo Labial")
        print("3) Liquidos")
        print("4) Tinta de Labios")
        print("5) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Lipstick")
            lipstick = obtener_lipsticks_en_categoria_lipstick()
            print(lipstick)
        elif elegirOpcion == "2":
            print("Brillo Labial")
            lip_gloss = obtener_lipsticks_en_categoria_lip_gloss()
            print(lip_gloss)
        elif elegirOpcion == "3":
            print("Liquidos")    
            liquid = obtener_lipsticks_en_categoria_liquid()
            print(liquid)
        elif elegirOpcion == "4":
            print("Tinta de Labios")  
            lip_stain = obtener_lipsticks_en_categoria_lip_stain()
            print(lip_stain)  
        elif elegirOpcion == "5":
            break
  #.......Menu Marcas de Lipstick.....
def MarcasLipstick():
    while True:
        print("\n---MARCAS LIPSTICK---")
        print("1) Informacion de marcas")
        print("2) Ver grafico")
        print("3) Salir")
        print("--------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Las Marcas de Lipstick son: ")
            marcasLipsticks = obtener_lipsticks_por_marca()
            print(marcasLipsticks)
        elif elegirOpcion == "2":
            print(" ver grafico: ")
            grafica_marcas_lipstick(marcas_unicas, cantidad_por_marca)
        elif elegirOpcion == "3":
            break

  #.......Menu Tonos de Lipstick.....
def TonosEyeshadow():
    while True:
        print("\n---TONOS DE LIPSTICK---")
        print("1) Tonos de marcas")
        print("2) Ver grafico")
        print("3) Salir")
        print("--------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Los Tonos de Lipstick son: ")
            tonosLipsticks = obtener_tonos_por_lipstick()
            print(tonosLipsticks) 
        elif elegirOpcion == "2":
            print(" ver grafico: ")
            grafico_tonos_lipstick()
        elif elegirOpcion == "3":
            break

  #.......Menu precios de Lipstick.....
def PrecioLipstick():
    while True:
        print("\n---PRECIO DE LIPSTICK---")
        print("1) Mas caros")
        print("2) Mas baratos")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            OpcionMasCarosL()
        elif elegirOpcion == "2":
            OpcionMasBaratosL()
        elif elegirOpcion == "3":
            break

    #________Menu Precios mas caros de Lipstick______
def OpcionMasCarosL():
    while True:
        print("\n---MAS CAROS---")
        print("1) Ver productos")
        print("2) Ver grafico")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            mostrar_lipsticks_mas_caros()
        elif elegirOpcion == "2":
            parq()
        elif elegirOpcion == "3":
            break

    #________Menu Precios mas baratos de Lipstick______
def OpcionMasBaratosL():
    while True:
        print("\n---MAS BARATOS---")
        print("1) Ver productos")
        print("2) Ver grafico")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            mostrar_productos_mas_baratos_lipstick()
        elif elegirOpcion == "2":
            mostrar_grafica_diferencia_precioL()
        elif elegirOpcion == "3":
            break

#//////////////////////// PRODUCTO MASCARA //////////////////////// 

#------- Se hacen los llamados y creaciones de datos para Etiquetas Mascara -------
  #.....seccion de obtencion de datos de precio para los productos Mascara.....

#....llamado a los precios mas caros de mascara....

def obtener_productos_mas_carosM(categories):
    expensive_products = []
    for category in categories:
        url = f"https://makeup-api.herokuapp.com/api/v1/products.json?product_tags={category}&product_type=mascara"
        response = requests.get(url)
        if response.status_code == 200:
            products = response.json()
            expensive_products.extend(products)
        else:
            print(f"Hubo un problema al obtener los productos de la categoria {category}.")
    return expensive_products

def obtener_mas_caros_mascara():
    categories = ["Natural", "Canadian", "Gluten+Free"]
    expensive_mascara = obtener_productos_mas_carosM(categories)
    most_expensive_mascara = []
    if expensive_mascara:
        expensive_mascara = [producto for producto in expensive_mascara if producto.get('price') is not None]
        most_expensive_mascara = sorted(expensive_mascara, key=lambda x: float(x.get('price', 0)), reverse=True)[:15]       
    return most_expensive_mascara

def mostrar_mascara_mas_cara():
    most_expensive_mascara = obtener_mas_caros_mascara()   
    if most_expensive_mascara:
        print("Los 15 productos mas caros de mascara son:")
        for index, producto in enumerate(most_expensive_mascara, 1):
            price = float(producto['price']) if producto['price'] else None
            print(f"{index}. {producto['name']} - Precio: {price}")

   #....Se crea la operacion matematica de los precios mas caros....
def operacion_matematica_preciosM():
    most_expensive_mascara = obtener_mas_caros_mascara()
    if most_expensive_mascara:
        precios = [float(producto['price']) for producto in most_expensive_mascara]
        suma_precios = sum(precios)       
        return suma_precios

def mostrar_mascara_mas_cara():
    most_expensive_mascara = obtener_mas_caros_mascara()    
    if most_expensive_mascara:
        print("Los 15 productos mas caros de mascara son:")
        for index, producto in enumerate(most_expensive_mascara, 1):
            price = float(producto['price']) if producto['price'] else None
            print(f"{index}. {producto['name']} - Precio: {price}")
        operacion_matematica = operacion_matematica_preciosM()
        if operacion_matematica:
            print(f"Resultado de la operacion matematica (suma de precios): {operacion_matematica}")

  #....Crear grafica para productos mas caros de mascara...
def mostrar_graficaM():
    most_expensive_mascara = obtener_mas_caros_mascara()  
    if most_expensive_mascara:
        productos = []
        precios = []       
        for index, producto in enumerate(most_expensive_mascara, 1):
            price = float(producto['price']) if producto['price'] else None
            productos.append(f"{index}. {producto['name']}")
            precios.append(price)
        suma_precios = sum(precios)
        productos.append("Suma de Precios")
        precios.append(suma_precios)
        plt.figure(figsize=(10, 6))
        plt.bar(productos, precios, color='skyblue')
        plt.xlabel('Productos')
        plt.ylabel('Precio')
        plt.title('Precios de los 15 productos mas caros de mascara')
        plt.xticks(rotation=90)
        plt.tight_layout()
        plt.show()

#....llamado a los precios mas baratos de mascara....

def obtener_productos_mas_baratosMM(categories):
    cheap_products = []
    for category in categories:
        url = f"https://makeup-api.herokuapp.com/api/v1/products.json?product_tags={category}&product_type=mascara"
        response = requests.get(url)
        if response.status_code == 200:
            products = response.json()
            cheap_products.extend(products)
        else:
            print(f"Hubo un problema al obtener los productos de la categoria {category}.")
    return cheap_products

def obtener_mas_baratos_mascara():
    categories = ["Natural", "Canadian", "Gluten+Free"]
    cheap_mascara = obtener_productos_mas_baratosMM(categories)
    cheapest_mascara = []
    if cheap_mascara:
        cheap_mascara = [producto for producto in cheap_mascara if producto.get('price') is not None]
        cheapest_mascara = sorted(cheap_mascara, key=lambda x: float(x.get('price', 0)))[:15]      
    return cheapest_mascara
   #......operacion matematica....
def calcular_serie_preciosM():
    cheapest_mascara = obtener_mas_baratos_mascara()
    if cheapest_mascara:
        precios = [float(producto['price']) for producto in cheapest_mascara]
        suma_precios = sum(precios)
        return suma_precios

def mostrar_mas_baratos_mascara_y_operacion():
    cheapest_mascara = obtener_mas_baratos_mascara()   
    if cheapest_mascara:
        print("Los 15 productos mas baratos de mascara son:")
        for index, producto in enumerate(cheapest_mascara, 1):
            price = float(producto['price']) if producto['price'] else None
            print(f"{index}. {producto['name']} - Precio: {price}")

        operacion_matematica = calcular_serie_preciosM()
        if operacion_matematica:
            print(f"Serie de los precios: {operacion_matematica}")

  #......grafico de los 15 productos mas baratos de mascara.....

def graficar_precios_y_productosM():
    cheapest_mascara = obtener_mas_baratos_mascara()

    if cheapest_mascara:
        precios = [float(producto['price']) for producto in cheapest_mascara]
        suma_precios = sum(precios)
        numeros_productos = list(range(1, len(cheapest_mascara) + 1))
        plt.figure(figsize=(8, 6))
        plt.plot(numeros_productos, precios, marker='o', linestyle='-', label='Precios')
        plt.axhline(y=suma_precios, color='r', linestyle='--', label='Suma de Precios')
        plt.xlabel('Productos')
        plt.ylabel('Precio')
        plt.title('Precios de los 15 productos mas baratos de mascara')
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.show()

#....grafico marcas mascara....
def grafica_etiquetas_mascara(etiquetas, conteos):
    plt.pie(conteos, labels=etiquetas, autopct='%1.1f%%')
    plt.axis('equal')  
    plt.title('Distribucion de etiquetas de Mascara')
    plt.show()

url = "https://makeup-api.herokuapp.com/api/v1/products.json"
params = {"product_type": "mascara"}
response = requests.get(url, params=params)

if response.status_code == 200:
    data = response.json()
    etiquetas = [producto['tag_list'] for producto in data]
    etiquetas = [etiqueta for sublist in etiquetas for etiqueta in sublist] 

    conteo_etiquetas = {}
    for etiqueta in etiquetas:
        conteo_etiquetas[etiqueta] = conteo_etiquetas.get(etiqueta, 0) + 1

    etiquetas_unicas = list(conteo_etiquetas.keys())
    cantidad_por_etiqueta = list(conteo_etiquetas.values())

#....grafico tonos mascara....
def grafico_tonos_por_mascara():
    url = "https://makeup-api.herokuapp.com/api/v1/products.json?product_type=mascara"

    try:
        response = requests.get(url)
        if response.status_code == 200:
            datos = response.json()
            tonos_por_mascara = {}
            for producto in datos:
                nombre = producto.get('name', 'Nombre no disponible')
                tonos = producto.get('product_colors', [])
                tonos_por_mascara[nombre] = tonos
            nombres_tonos = set()
            for tonos in tonos_por_mascara.values():
                for tono in tonos:
                    nombre_tono = tono.get('colour_name', 'Nombre no disponible')
                    nombres_tonos.add(nombre_tono)

            # creamos diccionario
            tonos_numerados = {nombre_tono: i + 1 for i, nombre_tono in enumerate(nombres_tonos)}
            nombres_mascara = []
            numeros_tonos = []

            for nombre, tonos in tonos_por_mascara.items():
                nombres_mascara.extend([nombre] * len(tonos))
                numeros_tonos.extend([tonos_numerados.get(t.get('colour_name')) for t in tonos])
            plt.figure(figsize=(10, 6))
            plt.scatter(nombres_mascara, numeros_tonos, color='grey', alpha=0.5)
            plt.xlabel('Mascaras')
            plt.ylabel('Cantidad de Tonos de Mascara')
            plt.title('Grafico de dispersion de tonos de Mascara')
            plt.xticks(rotation=90)
            plt.tight_layout()
            plt.show()
        else:
            print(f"Error al obtener datos. Codigo de estado: {response.status_code}")

    except requests.RequestException as e:
        print(f"Error de solicitud: {e}")

#-------------------Menu de Productos Mascara------------------
def Mascara():
  while True:
    print("\n---MENU DE MASCARA---")
    print("1) Etiquetas")
    print("2) Mejores calificaciones")
    print("3) precio")
    print("4) Marcas")
    print("5) Tonos")
    print("6) Regresar")
    print("----------------------------------------")
    elegirOpcion = input("elige una opcion: ")
    if elegirOpcion == "1":
      EtiquetaMascara()
    elif elegirOpcion == "2":
      print("Mejores calificaciones")
      mejoresMascaras = obtener_mascaras_con_calificacion_superior(4)
      print(mejoresMascaras)
    elif elegirOpcion == "3":
      PrecioMascara()
    elif elegirOpcion == "4":
      print("Marcas")
      MarcasMascara()
    elif elegirOpcion == "5":
      print("Tonos")
      TonosMascara()
    elif elegirOpcion == "6":
      break

  #.......Menu etiquetas de Mascara.....
def EtiquetaMascara():
    while True:
        print("\n---ETIQUETAS DE MASCARA---")
        print("1) Vegan")
        print("2) Hipoalergenicos")
        print("3) Organicos")
        print("4) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opción: ")
        if elegirOpcion == "1":
            print("Vegan")
            vegan = obtener_mascaras_vegan()
            print(vegan)
        elif elegirOpcion == "2":
            print("Hipoalergenicos")
            hypoallergenic = obtener_mascaras_hipoalergenicas()
            print(hypoallergenic)
        elif elegirOpcion == "3":
            print("Organicos")    
            organic = obtener_mascaras_organicas()
            print(organic)          
        elif elegirOpcion == "4":
            break

  #.......Menu Marcas de Mascara.....
def MarcasMascara():
    while True:
        print("\n---MARCAS MASCARA---")
        print("1) Informacion de marcas")
        print("2) Ver grafico")
        print("3) Salir")
        print("--------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Las Marcas de Mascara son: ")
            marcasMascaras = obtener_mascaras_por_marca()
            print(marcasMascaras)
        elif elegirOpcion == "2":
            print(" ver grafico: ")
            grafica_etiquetas_mascara(etiquetas_unicas, cantidad_por_etiqueta)
        elif elegirOpcion == "3":
            break

  #.......Menu Tonos de Mascara.....
def TonosMascara():
    while True:
        print("\n---TONOS DE MASCARA---")
        print("1) Tonos de marcas")
        print("2) Ver grafico")
        print("3) Salir")
        print("--------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            print("Los Tonos de Mascara son: ")
            tonosMascaras = obtener_tonos_por_mascara()
            print(tonosMascaras)
        elif elegirOpcion == "2":
            print(" ver grafico: ")
            grafico_tonos_por_mascara()
        elif elegirOpcion == "3":
            break

  #.......Menu precios de Mascara.....
def PrecioMascara():
    while True:
        print("\n---PRECIO DE MASCARA---")
        print("1) Mas caros")
        print("2) Mas baratos")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            OpcionMasCarosM()
        elif elegirOpcion == "2":
            OpcionMasBaratosM()
        elif elegirOpcion == "3":
            break

    #________Menu Precios mas caros de Mascara______
def OpcionMasCarosM():
    while True:
        print("\n---MAS CAROS---")
        print("1) Ver productos")
        print("2) Ver grafico")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            mostrar_mascara_mas_cara()
        elif elegirOpcion == "2":
            mostrar_graficaM()
        elif elegirOpcion == "3":
            break

    #________Menu Precios mas baratos de Mascara______
def OpcionMasBaratosM():
    while True:
        print("\n---MAS BARATOS---")
        print("1) Ver productos")
        print("2) Ver grafico")
        print("3) Regresar")
        print("----------------------------------------")
        elegirOpcion = input("elige una opcion: ")
        if elegirOpcion == "1":
            mostrar_mas_baratos_mascara_y_operacion()
        elif elegirOpcion == "2":
            graficar_precios_y_productosM()
        elif elegirOpcion == "3":
            break
#////////////////////reporte(excel)/////////////////////

def generar_reporte_productos():
    # Obtener datos de la API
    url = "http://makeup-api.herokuapp.com/api/v1/products.json"
    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()
        # Procesar los datos y seleccionar información relevante
        products_by_category = {}

        for product in data:
            category = product.get('category')
            name = product.get('name')

            if category not in products_by_category:
                products_by_category[category] = []

            products_by_category[category].append(name)

        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()

        # Crear una nueva hoja para listar los productos por categoría en columnas separadas
        sheet_products = wb.active
        sheet_products.title = "Productos por Categoría"

        # Escribir los productos por categoría en columnas separadas
        max_products = max(len(products) for products in products_by_category.values())

        # Escribir encabezados de categorías
        categories = list(products_by_category.keys())
        for col, category in enumerate(categories, start=1):
            sheet_products.cell(row=1, column=col).value = category

        # Escribir los productos en columnas correspondientes a cada categoría
        for col, category in enumerate(categories, start=1):
            products = products_by_category[category]
            for row, product in enumerate(products, start=2):
                sheet_products.cell(row=row, column=col).value = product

        # Guardar el archivo de Excel
        nombre_archivo = "reporte_productos_por_categoria.xlsx"
        wb.save(nombre_archivo)
        print(f"Se ha creado el archivo {nombre_archivo} con los datos de productos por categoría.")
    else:
        print("Hubo un problema al obtener los datos de la API.")


#/////////////////////encuesta/////////////////////

def realizar_encuesta():
    try:
        libro_excel = openpyxl.load_workbook('encuesta.xlsx')
    except FileNotFoundError:
        libro_excel = openpyxl.Workbook()
    hoja_encuesta = libro_excel.active
    if hoja_encuesta.title != "Encuesta":
        hoja_encuesta.title = "Encuesta"
    if not hoja_encuesta['A1'].value:
        encabezado = ["Pregunta 1", "Pregunta 2", "Pregunta 3", "Pregunta 4", "Correo Electrónico"]
        hoja_encuesta.append(encabezado)
    preguntas = [
        "¿Cual es tu nombre?",
        "¿Esta pagina fue de tu agrado? (a) Si (b) Algo (c) Normal (d) No",
        "¿Algun comentario que gustes agregar?",
        "¿Que calificacion le brindarias a nuestra pagina? (1-5 estrellas)"
    ]
    nueva_fila = []
    for pregunta in preguntas:
        respuesta = input(pregunta + " ")
        nueva_fila.append(respuesta)
    
    # Validación de datos para correo
    correo_valido = False
    while not correo_valido:
        correo = input("Ingresa tu correo electrónico para recibir notificaciones de ofertas: ")
        if re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', correo):
            nueva_fila.append(correo)
            correo_valido = True
        else:
            print("Correo electrónico inválido. Por favor, ingresa un correo válido.")

    hoja_encuesta.append(nueva_fila)

    libro_excel.save('encuesta.xlsx')
    print("¡Encuesta completada y registrada con éxito!")
    
#/////////////////////opiniones publicas(Estudio de Mercado)/////////////////////

def ResumenForms(Resumen):
    webbrowser.open(Resumen)

LinkResumen = "https://forms.office.com/Pages/AnalysisPage.aspx?AnalyzerToken=bbetmZ36dWSKl05m3MEBtr49nsMdSET6&id=EZDKymp73kSGHwlaLKiDt1FQrm9oY5hLjetXSxVOq6ZUMjRNTzEwQ0RCTFpUQUk1OU1CQlc2VkpaSi4u"


#/////////////////////creadoras del programa/////////////////////

def CreadorasOP():
  print("Seccion no.5 CreadorasOP")
  print(" ")
  print("Nombre: Samantha Elizabeth Salinas Pedraza")
  print("Matricula: 1901121")
  print(". . . . . . . . . . . . . . . . . . . . . . . .")
  print("Nombre: Kenya Jenling Vertiz Castro")
  print("Matricula: 1964884")
  print(". . . . . . . . . . . . . . . . . . . . . . . .")
  print("Nombre: Lesly Alejandra Rodriguez Alvarez")
  print("Matricula: 19034082")
  print(". . . . . . . . . . . . . . . . . . . . . . . .")
  
#///////////////////// MENU PRINCIPAL/////////////////////
 #...modulo sys...
print("Bienvenido a Makeup, por favor ingresa tu nombre:")
nombre = sys.stdin.readline().strip()
print(f"Bienvenido a Makeup, {nombre}")
  #...entrada a menu principal...
while True:
    print("\nMENÚ PRINCIPAL")
    print("1. Productos")
    print("2. Breve Encuesta(Excel)")
    print("3. Estudio de Mercado(Forms)")
    print("4. Creadoras del Codigo")
    print("5. Reporte de productos(excel)")
    print("6. Salir")
    print("----------------------------------------")

    opcion = input("seleccione una opcion: ")
    if opcion == "1":
        ProductosOP()  
    elif opcion == "2":
        realizar_encuesta()
    elif opcion == "3":
        ResumenForms(LinkResumen)
    elif opcion == "4": 
        CreadorasOP()
    elif opcion == "5":
        generar_reporte_productos()
    elif opcion == "6":
        print("salida exitosa")
        break
    else:
        print("ingrese opcion valida")

#/////////////////////CALIFICAR MENÚ/////////////////////
class Nodo:
    def __init__(self, categoria, calificacion):
        self.categoria= categoria
        self.calificacion= calificacion
        self.izquierda= None
        self.derecha= None
#.......................................................................................

def insertarNodo(raiz, categoria, calificacion):
    if raiz is None:
        return Nodo(categoria, calificacion)    
    if calificacion < raiz.calificacion:
        raiz.izquierda= insertarNodo(raiz.izquierda, categoria, calificacion)
    else:
        raiz.derecha= insertarNodo(raiz.derecha, categoria, calificacion)
    return raiz
#..........................................................................................................

def recorrerArbol_inorden(raiz):
    if raiz:
        recorrerArbol_inorden(raiz.izquierda)
        print(f"{raiz.categoria}: {raiz.calificacion}")
        recorrerArbol_inorden(raiz.derecha)

#...........................................................................................................

class Maquillaje:
    def __init__(self):
        self.calificaciones= {"PRODUCTOS": None, "ENCUESTAS": None, "ESTUDIO DE MERCADO": None, "GRÁFICOS": None, "REPORTE DE PRODUCTOS": None}

    def asignarCalificacion(self, categoria, calificacion):
        if categoria in self.calificaciones:
            if 0<= calificacion <= 100:
                self.calificaciones[categoria] = calificacion
                print(f"Calificación de {categoria} asignada: {calificacion}")
            else:
                print("La calificación debe estar en el rango de 0 a 100.")
        else:
            print("Categoría no válida. Las opciones son: Rubor, Labial, Sombras de ojos, Rimel")

    def ingresarCalificaciones(self):
        for categoria in self.calificaciones.keys():
            calificacion= int(input(f"Ingrese la calificación para {categoria} (entre 0 y 100): "))
            self.asignarCalificacion(categoria, calificacion)

# Ejemplo de uso
maquillaje= Maquillaje()
maquillaje.ingresarCalificaciones()
raiz_arbol= None
for categoria, calificacion in maquillaje.calificaciones.items():
    raiz_arbol= insertarNodo(raiz_arbol, categoria, calificacion)

print("\nCalificaciones ordenadas de menor a mayor: ")
recorrerArbol_inorden(raiz_arbol)